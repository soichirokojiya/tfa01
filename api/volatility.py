"""
Vercel Serverless Function: ボラティリティ計算Excel生成
POST /api/volatility

SPEEDAエクスポートの月次株価Excelを受け取り、
計算過程付きのボラティリティ計算Excelを返す。
"""

import io
import cgi
import tempfile
import os
import urllib.parse
from http.server import BaseHTTPRequestHandler
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


def generate_volatility_excel(input_bytes: bytes) -> bytes:
    wb_in = load_workbook(io.BytesIO(input_bytes))
    ws_in = wb_in.active

    data_rows = []
    for row in ws_in.iter_rows(min_row=2, values_only=True):
        if row[0] is not None and row[1] is not None:
            data_rows.append((row[0], row[1]))

    if len(data_rows) < 3:
        raise ValueError("月次株価データが不足しています（最低3ヶ月分必要）")

    n = len(data_rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "ボラティリティ計算"

    # スタイル定義
    title_font = Font(name="ＭＳ Ｐゴシック", size=14, bold=True)
    header_font = Font(name="ＭＳ Ｐゴシック", size=11, bold=True)
    data_font = Font(name="ＭＳ Ｐゴシック", size=11)
    formula_font = Font(name="ＭＳ Ｐゴシック", size=11, color="0000CC")
    result_font = Font(name="ＭＳ Ｐゴシック", size=12, bold=True, color="CC0000")
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    result_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # タイトル
    ws.merge_cells("A1:D1")
    ws["A1"] = "ボラティリティ算出シート"
    ws["A1"].font = title_font

    ws["A2"] = "※すべてのセルにExcel数式が入っており、計算過程を確認できます"
    ws["A2"].font = Font(name="ＭＳ Ｐゴシック", size=9, color="666666")

    # ヘッダー (行4)
    headers = [
        ("A", "No."),
        ("B", "年月"),
        ("C", "月次終値"),
        ("D", "対数収益率 ln(Pt/Pt-1)"),
    ]
    for col_letter, label in headers:
        cell = ws[f"{col_letter}4"]
        cell.value = label
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 28

    # データ行 (行5～)
    data_start_row = 5
    for i, (date_val, price_val) in enumerate(data_rows):
        row_num = data_start_row + i

        cell_a = ws[f"A{row_num}"]
        cell_a.value = i + 1
        cell_a.font = data_font
        cell_a.border = thin_border
        cell_a.alignment = Alignment(horizontal="center")

        cell_b = ws[f"B{row_num}"]
        cell_b.value = date_val
        cell_b.font = data_font
        cell_b.border = thin_border
        if isinstance(date_val, str):
            cell_b.alignment = Alignment(horizontal="center")
        else:
            cell_b.number_format = "YYYY/MM"
            cell_b.alignment = Alignment(horizontal="center")

        cell_c = ws[f"C{row_num}"]
        cell_c.value = price_val
        cell_c.font = data_font
        cell_c.border = thin_border
        cell_c.number_format = "#,##0"

        cell_d = ws[f"D{row_num}"]
        if i == 0:
            cell_d.value = "―"
            cell_d.font = data_font
            cell_d.alignment = Alignment(horizontal="center")
        else:
            cell_d.value = f"=LN(C{row_num}/C{row_num - 1})"
            cell_d.font = formula_font
            cell_d.number_format = "0.000000"
        cell_d.border = thin_border

    data_end_row = data_start_row + n - 1
    return_start_row = data_start_row + 1
    return_range = f"D{return_start_row}:D{data_end_row}"

    # 計算結果セクション
    calc_start = data_end_row + 2

    ws.merge_cells(f"A{calc_start}:D{calc_start}")
    ws[f"A{calc_start}"] = "【ボラティリティ計算】"
    ws[f"A{calc_start}"].font = header_font

    # (1) 月次標準偏差
    r1 = calc_start + 1
    ws[f"A{r1}"] = "①"
    ws[f"A{r1}"].font = data_font
    ws[f"A{r1}"].alignment = Alignment(horizontal="center")
    ws[f"B{r1}"] = "月次σ（標準偏差）"
    ws[f"B{r1}"].font = data_font
    ws.merge_cells(f"B{r1}:C{r1}")
    cell_monthly = ws[f"D{r1}"]
    cell_monthly.value = f"=STDEV({return_range})"
    cell_monthly.font = result_font
    cell_monthly.number_format = "0.000000"
    cell_monthly.fill = result_fill
    cell_monthly.border = thin_border

    r1e = r1 + 1
    ws[f"B{r1e}"] = f"  = STDEV({return_range})"
    ws[f"B{r1e}"].font = Font(name="ＭＳ Ｐゴシック", size=9, color="666666")

    # (2) 年率換算
    r2 = r1e + 1
    ws[f"A{r2}"] = "②"
    ws[f"A{r2}"].font = data_font
    ws[f"A{r2}"].alignment = Alignment(horizontal="center")
    ws[f"B{r2}"] = "年率σ（年率換算）"
    ws[f"B{r2}"].font = data_font
    ws.merge_cells(f"B{r2}:C{r2}")
    cell_annual = ws[f"D{r2}"]
    cell_annual.value = f"=D{r1}*SQRT(12)"
    cell_annual.font = result_font
    cell_annual.number_format = "0.000000"
    cell_annual.fill = result_fill
    cell_annual.border = thin_border

    r2e = r2 + 1
    ws[f"B{r2e}"] = f"  = D{r1} × √12  （月次 → 年率換算）"
    ws[f"B{r2e}"].font = Font(name="ＭＳ Ｐゴシック", size=9, color="666666")

    # (3) パーセント表示
    r3 = r2e + 1
    ws[f"A{r3}"] = "③"
    ws[f"A{r3}"].font = data_font
    ws[f"A{r3}"].alignment = Alignment(horizontal="center")
    ws[f"B{r3}"] = "年率σ（%表示）"
    ws[f"B{r3}"].font = data_font
    ws.merge_cells(f"B{r3}:C{r3}")
    cell_pct = ws[f"D{r3}"]
    cell_pct.value = f"=D{r2}*100"
    cell_pct.font = result_font
    cell_pct.number_format = '0.00"%"'
    cell_pct.fill = result_fill
    cell_pct.border = thin_border

    r3e = r3 + 1
    ws[f"B{r3e}"] = f"  = D{r2} × 100"
    ws[f"B{r3e}"].font = Font(name="ＭＳ Ｐゴシック", size=9, color="666666")

    # 算出方法の説明
    note_start = r3e + 2
    ws[f"A{note_start}"] = "【算出方法】"
    ws[f"A{note_start}"].font = header_font

    notes = [
        "1. 月次株価の終値データを使用",
        "2. 対数収益率 = LN(当月終値 / 前月終値) を各月について算出",
        "3. 対数収益率の標準偏差（STDEV）を月次ボラティリティとする",
        "4. 年率ボラティリティ = 月次ボラティリティ × √12",
    ]
    for j, note in enumerate(notes):
        ws[f"A{note_start + 1 + j}"] = note
        ws[f"A{note_start + 1 + j}"].font = Font(name="ＭＳ Ｐゴシック", size=10)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class handler(BaseHTTPRequestHandler):
    def do_POST(self):
        try:
            content_type = self.headers.get("Content-Type", "")
            content_length = int(self.headers.get("Content-Length", 0))

            if "multipart/form-data" in content_type:
                # multipart/form-data からファイルを抽出
                environ = {
                    "REQUEST_METHOD": "POST",
                    "CONTENT_TYPE": content_type,
                    "CONTENT_LENGTH": str(content_length),
                }
                form = cgi.FieldStorage(
                    fp=self.rfile,
                    headers=self.headers,
                    environ=environ,
                )
                file_item = form["file"]
                file_bytes = file_item.file.read()
            else:
                file_bytes = self.rfile.read(content_length)

            result_bytes = generate_volatility_excel(file_bytes)
            filename = "ボラティリティ計算.xlsx"

            self.send_response(200)
            self.send_header("Content-Type",
                             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition",
                             f"attachment; filename*=UTF-8''{urllib.parse.quote(filename)}")
            self.send_header("Content-Length", str(len(result_bytes)))
            self.end_headers()
            self.wfile.write(result_bytes)

        except Exception as e:
            import json
            self.send_response(500)
            self.send_header("Content-Type", "application/json")
            self.end_headers()
            self.wfile.write(json.dumps({"error": str(e)}).encode())
