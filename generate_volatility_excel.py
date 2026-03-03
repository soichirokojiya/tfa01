#!/usr/bin/env python3
"""
ボラティリティ計算Excel生成スクリプト

SPEEDAからエクスポートした月次株価データ（Excel）を読み込み、
計算過程がすべてExcel数式で見えるボラティリティ計算シートを生成する。

使い方:
  python3 generate_volatility_excel.py <入力Excelファイル>

入力Excelの想定フォーマット:
  - 1行目: ヘッダー（日付, 終値 など）
  - A列: 日付（月次）
  - B列: 終値（月次株価）
"""

import sys
import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter

OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")


def generate_volatility_excel(input_path: str, output_path: str = None):
    """
    入力Excelの月次株価から、計算過程付きのボラティリティ計算Excelを生成。
    すべてExcel数式で記述し、計算ロジックが透明に見えるようにする。
    """

    # ── 入力ファイル読み込み ──
    wb_in = load_workbook(input_path)
    ws_in = wb_in.active

    # データ読み取り（日付と終値）
    data_rows = []
    for row in ws_in.iter_rows(min_row=2, values_only=True):
        if row[0] is not None and row[1] is not None:
            # 数値データ行のみ取り込む（SPEEDAのヘッダー行をスキップ）
            try:
                float(row[1])
            except (ValueError, TypeError):
                continue
            data_rows.append((row[0], row[1]))

    if len(data_rows) < 3:
        raise ValueError("月次株価データが不足しています（最低3ヶ月分必要）")

    n = len(data_rows)

    # ── 出力Excelを作成 ──
    wb = Workbook()
    ws = wb.active
    ws.title = "ボラティリティ計算"

    # スタイル定義
    title_font = Font(name="ＭＳ Ｐゴシック", size=14, bold=True)
    header_font = Font(name="ＭＳ Ｐゴシック", size=11, bold=True)
    data_font = Font(name="ＭＳ Ｐゴシック", size=11)
    formula_font = Font(name="ＭＳ Ｐゴシック", size=11, color="0000CC")
    result_font = Font(name="ＭＳ Ｐゴシック", size=12, bold=True, color="CC0000")
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    result_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # ── タイトル ──
    ws.merge_cells("A1:D1")
    ws["A1"] = "ボラティリティ算出シート"
    ws["A1"].font = title_font

    ws["A2"] = "※すべてのセルにExcel数式が入っており、計算過程を確認できます"
    ws["A2"].font = Font(name="ＭＳ Ｐゴシック", size=9, color="666666")

    # ── ヘッダー (行4) ──
    headers = [
        ("A", "No."),
        ("B", "年月"),
        ("C", "月次終値"),
        ("D", "対数収益率 ln(Pt/Pt-1)"),
    ]
    for col_letter, label in headers:
        cell = ws[f"{col_letter}4"]
        cell.value = label
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # 列幅設定
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 28

    # ── データ行 (行5～) ──
    data_start_row = 5
    for i, (date_val, price_val) in enumerate(data_rows):
        row_num = data_start_row + i

        # No.
        cell_a = ws[f"A{row_num}"]
        cell_a.value = i + 1
        cell_a.font = data_font
        cell_a.border = thin_border
        cell_a.alignment = Alignment(horizontal="center")

        # 年月
        cell_b = ws[f"B{row_num}"]
        cell_b.value = date_val
        cell_b.font = data_font
        cell_b.border = thin_border
        if isinstance(date_val, str):
            cell_b.alignment = Alignment(horizontal="center")
        else:
            cell_b.number_format = "YYYY/MM"
            cell_b.alignment = Alignment(horizontal="center")

        # 月次終値
        cell_c = ws[f"C{row_num}"]
        cell_c.value = price_val
        cell_c.font = data_font
        cell_c.border = thin_border
        cell_c.number_format = "#,##0"

        # 対数収益率 = LN(Ct/Ct-1) ← Excel数式
        cell_d = ws[f"D{row_num}"]
        if i == 0:
            cell_d.value = "―"
            cell_d.font = data_font
            cell_d.alignment = Alignment(horizontal="center")
        else:
            cell_d.value = f"=LN(C{row_num}/C{row_num - 1})"
            cell_d.font = formula_font
            cell_d.number_format = "0.000000"
        cell_d.border = thin_border

    data_end_row = data_start_row + n - 1
    # 対数収益率の範囲 (最初の行は "―" なので除外)
    return_start_row = data_start_row + 1
    return_range = f"D{return_start_row}:D{data_end_row}"

    # ── 計算結果セクション ──
    calc_start = data_end_row + 2

    # 区切り線
    ws.merge_cells(f"A{calc_start}:D{calc_start}")
    ws[f"A{calc_start}"] = "【ボラティリティ計算】"
    ws[f"A{calc_start}"].font = header_font

    # (1) 月次対数収益率の標準偏差
    r1 = calc_start + 1
    ws[f"A{r1}"] = "①"
    ws[f"A{r1}"].font = data_font
    ws[f"A{r1}"].alignment = Alignment(horizontal="center")
    ws[f"B{r1}"] = "月次σ（標準偏差）"
    ws[f"B{r1}"].font = data_font
    ws.merge_cells(f"B{r1}:C{r1}")
    cell_monthly = ws[f"D{r1}"]
    cell_monthly.value = f"=STDEVP({return_range})"
    cell_monthly.font = result_font
    cell_monthly.number_format = "0.000000"
    cell_monthly.fill = result_fill
    cell_monthly.border = thin_border

    # 数式説明
    r1e = r1 + 1
    ws[f"B{r1e}"] = f"  = STDEVP({return_range})"
    ws[f"B{r1e}"].font = Font(name="ＭＳ Ｐゴシック", size=9, color="666666")

    # (2) 年率換算 = 月次σ × √12
    r2 = r1e + 1
    ws[f"A{r2}"] = "②"
    ws[f"A{r2}"].font = data_font
    ws[f"A{r2}"].alignment = Alignment(horizontal="center")
    ws[f"B{r2}"] = "年率σ（年率換算）"
    ws[f"B{r2}"].font = data_font
    ws.merge_cells(f"B{r2}:C{r2}")
    cell_annual = ws[f"D{r2}"]
    cell_annual.value = f"=D{r1}*SQRT(12)"
    cell_annual.font = result_font
    cell_annual.number_format = "0.000000"
    cell_annual.fill = result_fill
    cell_annual.border = thin_border

    # 数式説明
    r2e = r2 + 1
    ws[f"B{r2e}"] = f"  = D{r1} × √12  （月次 → 年率換算）"
    ws[f"B{r2e}"].font = Font(name="ＭＳ Ｐゴシック", size=9, color="666666")

    # (3) パーセント表示
    r3 = r2e + 1
    ws[f"A{r3}"] = "③"
    ws[f"A{r3}"].font = data_font
    ws[f"A{r3}"].alignment = Alignment(horizontal="center")
    ws[f"B{r3}"] = "年率σ（%表示）"
    ws[f"B{r3}"].font = data_font
    ws.merge_cells(f"B{r3}:C{r3}")
    cell_pct = ws[f"D{r3}"]
    cell_pct.value = f"=D{r2}*100"
    cell_pct.font = result_font
    cell_pct.number_format = '0.00"%"'
    cell_pct.fill = result_fill
    cell_pct.border = thin_border

    # 数式説明
    r3e = r3 + 1
    ws[f"B{r3e}"] = f"  = D{r2} × 100"
    ws[f"B{r3e}"].font = Font(name="ＭＳ Ｐゴシック", size=9, color="666666")

    # ── 計算方法の説明 ──
    note_start = r3e + 2
    ws[f"A{note_start}"] = "【算出方法】"
    ws[f"A{note_start}"].font = header_font

    notes = [
        "1. 月次株価の終値データを使用",
        "2. 対数収益率 = LN(当月終値 / 前月終値) を各月について算出",
        "3. 対数収益率の標準偏差（STDEV）を月次ボラティリティとする",
        "4. 年率ボラティリティ = 月次ボラティリティ × √12",
    ]
    for j, note in enumerate(notes):
        ws[f"A{note_start + 1 + j}"] = note
        ws[f"A{note_start + 1 + j}"].font = Font(name="ＭＳ Ｐゴシック", size=10)

    # ── 保存 ──
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    if output_path is None:
        base = os.path.splitext(os.path.basename(input_path))[0]
        output_path = os.path.join(OUTPUT_DIR, f"{base}_ボラティリティ計算.xlsx")

    wb.save(output_path)
    print(f"生成完了: {output_path}")
    return output_path


# ── デモ用: yfinanceデータでサンプル生成 ──
def generate_demo(ticker_code: str, eval_date: str):
    """yfinanceからデータ取得してデモ用Excelを生成"""
    import yfinance as yf
    from datetime import datetime, timedelta
    from dateutil.relativedelta import relativedelta
    import warnings
    warnings.filterwarnings("ignore")

    eval_dt = datetime.strptime(eval_date, "%Y-%m-%d")
    vol_end_month = eval_dt.replace(day=1) - timedelta(days=1)
    vol_start_month = vol_end_month - relativedelta(years=5)
    vol_start = vol_start_month.replace(day=1).strftime("%Y-%m-%d")
    vol_end = (vol_end_month + timedelta(days=1)).strftime("%Y-%m-%d")

    ticker = yf.Ticker(f"{ticker_code}.T")
    hist = ticker.history(start=vol_start, end=vol_end, interval="1mo")

    # 一時的な入力Excel生成
    wb_tmp = Workbook()
    ws_tmp = wb_tmp.active
    ws_tmp["A1"] = "年月"
    ws_tmp["B1"] = "終値"
    for i, (date, row) in enumerate(hist.iterrows()):
        ws_tmp[f"A{i+2}"] = date.strftime("%Y/%m")
        ws_tmp[f"B{i+2}"] = round(row["Close"])

    tmp_path = os.path.join(OUTPUT_DIR, "_tmp_input.xlsx")
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    wb_tmp.save(tmp_path)

    output_path = os.path.join(OUTPUT_DIR, f"{ticker_code}_ボラティリティ計算.xlsx")
    generate_volatility_excel(tmp_path, output_path)
    os.remove(tmp_path)


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("使い方:")
        print("  python3 generate_volatility_excel.py <SPEEDAエクスポート.xlsx>")
        print()
        print("デモ (yfinanceデータ):")
        print("  python3 generate_volatility_excel.py --demo 3070 2025-06-13")
        sys.exit(1)

    if sys.argv[1] == "--demo":
        if len(sys.argv) < 4:
            print("デモ: python3 generate_volatility_excel.py --demo <銘柄コード> <評価基準日>")
            sys.exit(1)
        generate_demo(sys.argv[2], sys.argv[3])
    else:
        generate_volatility_excel(sys.argv[1])
